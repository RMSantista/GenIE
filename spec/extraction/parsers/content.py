"""Content-to-text extraction helpers shared by the Connector agent.

Turns heterogeneous payloads (PDF, XLSX, CSV, JSON, HTML, plain text)
into plain text suitable for LLM extraction.
"""

import csv
import io
import json
import logging
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Optional

from PyPDF2 import PdfReader

from spec.core.exceptions import ExtractionFailed

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = frozenset(
    {
        ".pdf",
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".html",
        ".htm",
        ".xml",
        ".yaml",
        ".yml",
        ".xlsx",
        ".log",
        ".tsv",
    }
)

_MAX_TEXT_CHARS = 400_000


class _TextExtractor(HTMLParser):
    """Minimal HTML-to-text converter (keeps visible text only)."""

    _SKIP_TAGS = frozenset({"script", "style", "noscript", "template", "head"})

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._chunks: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list) -> None:
        if tag in self._SKIP_TAGS:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in self._SKIP_TAGS and self._skip_depth > 0:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth == 0 and data.strip():
            self._chunks.append(data.strip())

    def text(self) -> str:
        return "\n".join(self._chunks)


def html_to_text(html: str) -> str:
    """Convert HTML markup into readable plain text.

    Args:
        html: Raw HTML

    Returns:
        str: Visible text content
    """

    parser = _TextExtractor()
    try:
        parser.feed(html)
        return parser.text()
    except Exception:
        return re.sub(r"<[^>]+>", " ", html)


def pdf_bytes_to_text(data: bytes, name: str = "document.pdf") -> str:
    """Extract text from PDF bytes.

    Args:
        data: PDF file content
        name: Filename used in error messages

    Returns:
        str: Extracted text

    Raises:
        ExtractionFailed: If the PDF has no extractable text (e.g. scanned)
    """

    try:
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or "")
            except Exception as e:  # pragma: no cover - defensive per page
                logger.warning("Failed to read a page of %s: %s", name, e)
        text = "\n".join(p for p in parts if p).strip()
    except ExtractionFailed:
        raise
    except Exception as e:
        raise ExtractionFailed(f"Falha ao processar PDF '{name}': {e}")

    if not text:
        raise ExtractionFailed(
            f"PDF '{name}' não contém texto extraível (provavelmente digitalizado). "
            "Suporte a OCR está planejado; envie um PDF nativo ou um TXT/CSV."
        )
    return text


def xlsx_bytes_to_text(data: bytes, name: str = "sheet.xlsx") -> str:
    """Extract cell values from XLSX bytes as TSV-like text.

    Args:
        data: Workbook content
        name: Filename used in error messages

    Returns:
        str: Tab-separated rows, one sheet after another
    """

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Planilha: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        return "\n".join(lines)
    except ImportError:
        raise ExtractionFailed("openpyxl não instalado — necessário para ler .xlsx")
    except Exception as e:
        raise ExtractionFailed(f"Falha ao processar planilha '{name}': {e}")


def bytes_to_text(data: bytes, name: str, content_type: Optional[str] = None) -> str:
    """Convert raw bytes into text based on file extension / content type.

    Args:
        data: Raw payload
        name: Filename or URL (used to infer the format)
        content_type: Optional MIME type hint

    Returns:
        str: Plain text (truncated to a safe maximum)

    Raises:
        ExtractionFailed: If the format cannot be converted
    """

    suffix = Path(name.split("?")[0]).suffix.lower()
    ctype = (content_type or "").split(";")[0].strip().lower()

    if suffix == ".pdf" or ctype == "application/pdf":
        text = pdf_bytes_to_text(data, name)
    elif suffix == ".xlsx" or ctype.endswith("spreadsheetml.sheet"):
        text = xlsx_bytes_to_text(data, name)
    elif suffix in (".html", ".htm") or ctype == "text/html":
        text = html_to_text(data.decode("utf-8", errors="replace"))
    elif suffix == ".json" or ctype == "application/json":
        raw = data.decode("utf-8", errors="replace")
        try:
            text = json.dumps(json.loads(raw), indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            text = raw
    elif suffix in (".csv", ".tsv"):
        raw = data.decode("utf-8", errors="replace")
        delimiter = "\t" if suffix == ".tsv" else ","
        rows = list(csv.reader(io.StringIO(raw), delimiter=delimiter))
        text = "\n".join("\t".join(row) for row in rows)
    elif suffix == ".docx":
        raise ExtractionFailed(
            f"Formato .docx ainda não suportado ('{name}'). Converta para PDF ou TXT."
        )
    else:
        text = data.decode("utf-8", errors="replace")

    text = text.strip()
    if len(text) > _MAX_TEXT_CHARS:
        logger.warning(
            "Truncating '%s' from %d to %d chars", name, len(text), _MAX_TEXT_CHARS
        )
        text = text[:_MAX_TEXT_CHARS]
    return text


def file_to_text(path: Path) -> str:
    """Read a local file and convert it to plain text.

    Args:
        path: File path

    Returns:
        str: Plain text content
    """

    return bytes_to_text(path.read_bytes(), path.name)
